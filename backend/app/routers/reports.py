"""Report generation: Excel + PDF."""
from __future__ import annotations

import io
from datetime import datetime, timezone
from typing import Annotated, Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from sqlalchemy.orm import Session, joinedload

from app.auth import require_any
from app.database import get_db
from app.models import Contractor, Driver, User, Violation, ViolationType

router = APIRouter(prefix="/api/reports", tags=["reports"])


def _stream_xlsx(wb: Workbook, filename: str) -> StreamingResponse:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _header_style(ws, row: int, end_col: int):
    fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    font = Font(bold=True, color="FFFFFF")
    for col in range(1, end_col + 1):
        c = ws.cell(row=row, column=col)
        c.fill = fill
        c.font = font
        c.alignment = Alignment(horizontal="center", vertical="center")


@router.get("/violations.xlsx")
def export_violations_xlsx(
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[User, Depends(require_any)],
    driver_id: Optional[int] = None,
    contractor_id: Optional[int] = None,
    date_from: Optional[datetime] = None,
    date_to: Optional[datetime] = None,
):
    q = (
        db.query(Violation)
        .options(joinedload(Violation.driver).joinedload(Driver.contractor))
        .options(joinedload(Violation.violation_type))
    )
    if driver_id:
        q = q.filter(Violation.driver_id == driver_id)
    if contractor_id:
        q = q.join(Driver, Driver.id == Violation.driver_id).filter(
            Driver.contractor_id == contractor_id
        )
    if date_from:
        q = q.filter(Violation.occurred_at >= date_from)
    if date_to:
        q = q.filter(Violation.occurred_at <= date_to)
    rows = q.order_by(Violation.occurred_at.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Violations"
    headers = [
        "ID",
        "Date/Time",
        "Driver",
        "Contractor",
        "Violation Type",
        "Severity",
        "Location",
        "Vehicle Plate",
        "Speed (km/h)",
        "Reported By",
        "Description",
        "Triggered Status",
    ]
    ws.append(headers)
    _header_style(ws, 1, len(headers))
    for v in rows:
        ws.append([
            v.id,
            v.occurred_at.replace(tzinfo=None) if v.occurred_at else "",
            v.driver.full_name if v.driver else "",
            v.driver.contractor.name if v.driver and v.driver.contractor else "",
            v.violation_type.name_en if v.violation_type else "",
            v.severity.value if v.severity else "",
            v.location or "",
            v.vehicle_plate or "",
            v.speed_kmh if v.speed_kmh is not None else "",
            v.reported_by or "",
            v.description or "",
            v.triggered_status.value if v.triggered_status else "",
        ])
    for col_idx, _ in enumerate(headers, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 20

    return _stream_xlsx(wb, f"violations_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")


@router.get("/drivers.xlsx")
def export_drivers_xlsx(
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[User, Depends(require_any)],
):
    rows = db.query(Driver).options(joinedload(Driver.contractor)).order_by(Driver.full_name).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Drivers"
    headers = [
        "ID", "Full Name", "Full Name (AR)", "National ID", "License #",
        "License Expiry", "Defensive Driving Expiry", "Nationality", "Phone",
        "Contractor", "Status", "Blacklisted", "Active", "Notes",
    ]
    ws.append(headers)
    _header_style(ws, 1, len(headers))
    for d in rows:
        from sqlalchemy import func as _f
        count = db.query(_f.count(Violation.id)).filter(Violation.driver_id == d.id).scalar() or 0
        ws.append([
            d.id, d.full_name, d.full_name_ar or "", d.national_id or "", d.license_number or "",
            d.license_expiry.replace(tzinfo=None) if d.license_expiry else "",
            d.defensive_driving_expiry.replace(tzinfo=None) if d.defensive_driving_expiry else "",
            d.nationality or "", d.phone or "",
            d.contractor.name if d.contractor else "",
            d.status.value, "Yes" if d.is_blacklisted else "No",
            "Yes" if d.is_active else "No", d.notes or "",
        ])
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 20
    return _stream_xlsx(wb, f"drivers_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")


@router.get("/contractors.xlsx")
def export_contractors_xlsx(
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[User, Depends(require_any)],
):
    from sqlalchemy import func as _f
    rows = db.query(Contractor).order_by(Contractor.name).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Contractors"
    headers = [
        "ID", "Name", "Name (AR)", "Contact Person", "Phone", "Email",
        "Drivers", "Violating Drivers", "Total Violations", "Blacklisted", "Active",
    ]
    ws.append(headers)
    _header_style(ws, 1, len(headers))
    for c in rows:
        d_count = db.query(_f.count(Driver.id)).filter(Driver.contractor_id == c.id).scalar() or 0
        v_drv = (
            db.query(_f.count(Driver.id.distinct()))
            .join(Violation, Violation.driver_id == Driver.id)
            .filter(Driver.contractor_id == c.id)
            .scalar() or 0
        )
        total_v = (
            db.query(_f.count(Violation.id))
            .join(Driver, Driver.id == Violation.driver_id)
            .filter(Driver.contractor_id == c.id)
            .scalar() or 0
        )
        ws.append([
            c.id, c.name, c.name_ar or "", c.contact_person or "",
            c.contact_phone or "", c.contact_email or "",
            int(d_count), int(v_drv), int(total_v),
            "Yes" if c.is_blacklisted else "No",
            "Yes" if c.is_active else "No",
        ])
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 20
    return _stream_xlsx(wb, f"contractors_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")


@router.get("/driver/{driver_id}.pdf")
def driver_pdf(
    driver_id: int,
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[User, Depends(require_any)],
):
    d = (
        db.query(Driver)
        .options(joinedload(Driver.contractor))
        .filter(Driver.id == driver_id)
        .first()
    )
    if not d:
        raise HTTPException(status_code=404, detail="Driver not found")
    violations = (
        db.query(Violation)
        .options(joinedload(Violation.violation_type))
        .filter(Violation.driver_id == driver_id)
        .order_by(Violation.occurred_at.desc())
        .all()
    )

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=15 * mm, rightMargin=15 * mm,
                            topMargin=15 * mm, bottomMargin=15 * mm)
    styles = getSampleStyleSheet()
    title = ParagraphStyle("title", parent=styles["Heading1"], alignment=TA_CENTER,
                           textColor=colors.HexColor("#1F2937"))
    sub = ParagraphStyle("sub", parent=styles["Heading3"], textColor=colors.HexColor("#374151"))
    normal = styles["Normal"]

    story = []
    story.append(Paragraph("West Qurna Oil Field - Driver Safety Report", title))
    story.append(Paragraph("حقل غرب القرنة النفطي - تقرير سلامة السائق", title))
    story.append(Spacer(1, 8 * mm))

    info_data = [
        ["Name / الاسم", d.full_name + (f"  ({d.full_name_ar})" if d.full_name_ar else "")],
        ["National ID / الهوية الوطنية", d.national_id or "-"],
        ["License #", d.license_number or "-"],
        ["Nationality / الجنسية", d.nationality or "-"],
        ["Phone / الهاتف", d.phone or "-"],
        ["Contractor / المقاول", d.contractor.name if d.contractor else "-"],
        ["Status / الحالة", d.status.value.upper()],
        ["Blacklisted / محظور", "YES" if d.is_blacklisted else "NO"],
        ["Total Violations / مجموع المخالفات", str(len(violations))],
        ["Report Date", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")],
    ]
    t = Table(info_data, colWidths=[60 * mm, 110 * mm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#E5E7EB")),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#9CA3AF")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(t)
    story.append(Spacer(1, 8 * mm))

    story.append(Paragraph("Violations History / سجل المخالفات", sub))
    story.append(Spacer(1, 3 * mm))

    if not violations:
        story.append(Paragraph("No violations recorded.", normal))
    else:
        header = ["#", "Date", "Type", "Severity", "Location", "Plate", "Action"]
        rows = [header]
        for i, v in enumerate(violations, start=1):
            rows.append([
                str(i),
                v.occurred_at.strftime("%Y-%m-%d %H:%M") if v.occurred_at else "",
                v.violation_type.name_en if v.violation_type else "",
                v.severity.value if v.severity else "",
                (v.location or "")[:30],
                v.vehicle_plate or "",
                v.triggered_status.value if v.triggered_status else "",
            ])
        vt = Table(rows, repeatRows=1,
                   colWidths=[10 * mm, 30 * mm, 35 * mm, 18 * mm, 35 * mm, 22 * mm, 25 * mm])
        vt.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2937")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#9CA3AF")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F4F6")]),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        story.append(vt)

    story.append(Spacer(1, 10 * mm))
    story.append(Paragraph(
        "Issued by the Safety Department - West Qurna Oil Field<br/>"
        "صادر عن قسم السلامة - حقل غرب القرنة النفطي",
        ParagraphStyle("footer", parent=normal, alignment=TA_CENTER,
                       textColor=colors.HexColor("#6B7280"), fontSize=9),
    ))

    doc.build(story)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="driver_{driver_id}_report.pdf"'},
    )


@router.get("/contractor/{contractor_id}.pdf")
def contractor_pdf(
    contractor_id: int,
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[User, Depends(require_any)],
):
    from sqlalchemy import func as _f
    c = db.query(Contractor).filter(Contractor.id == contractor_id).first()
    if not c:
        raise HTTPException(status_code=404, detail="Contractor not found")
    drivers = db.query(Driver).filter(Driver.contractor_id == contractor_id).all()
    total_v = (
        db.query(_f.count(Violation.id))
        .join(Driver, Driver.id == Violation.driver_id)
        .filter(Driver.contractor_id == contractor_id)
        .scalar() or 0
    )

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=15 * mm, rightMargin=15 * mm,
                            topMargin=15 * mm, bottomMargin=15 * mm)
    styles = getSampleStyleSheet()
    title = ParagraphStyle("title", parent=styles["Heading1"], alignment=TA_CENTER,
                           textColor=colors.HexColor("#1F2937"))
    sub = ParagraphStyle("sub", parent=styles["Heading3"], textColor=colors.HexColor("#374151"))
    normal = styles["Normal"]
    story = [
        Paragraph("West Qurna Oil Field - Contractor Safety Report", title),
        Paragraph("حقل غرب القرنة النفطي - تقرير سلامة المقاول", title),
        Spacer(1, 8 * mm),
    ]

    info = [
        ["Contractor / المقاول", c.name + (f"  ({c.name_ar})" if c.name_ar else "")],
        ["Contact Person", c.contact_person or "-"],
        ["Phone", c.contact_phone or "-"],
        ["Email", c.contact_email or "-"],
        ["Total Drivers", str(len(drivers))],
        ["Total Violations", str(int(total_v))],
        ["Status", "BLACKLISTED" if c.is_blacklisted else ("ACTIVE" if c.is_active else "INACTIVE")],
        ["Report Date", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")],
    ]
    t = Table(info, colWidths=[60 * mm, 110 * mm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#E5E7EB")),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#9CA3AF")),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
    ]))
    story.append(t)
    story.append(Spacer(1, 8 * mm))

    story.append(Paragraph("Drivers Under This Contractor / السائقون التابعون", sub))
    story.append(Spacer(1, 3 * mm))

    rows = [["#", "Driver", "License #", "Status", "Violations"]]
    for i, d in enumerate(drivers, start=1):
        v_count = db.query(_f.count(Violation.id)).filter(Violation.driver_id == d.id).scalar() or 0
        rows.append([str(i), d.full_name, d.license_number or "-", d.status.value, str(int(v_count))])
    dt = Table(rows, repeatRows=1, colWidths=[10 * mm, 70 * mm, 35 * mm, 30 * mm, 25 * mm])
    dt.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2937")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#9CA3AF")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F4F6")]),
    ]))
    story.append(dt)
    story.append(Spacer(1, 8 * mm))
    story.append(Paragraph(
        "Issued by the Safety Department - West Qurna Oil Field",
        ParagraphStyle("footer", parent=normal, alignment=TA_CENTER,
                       textColor=colors.HexColor("#6B7280"), fontSize=9),
    ))
    doc.build(story)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="contractor_{contractor_id}_report.pdf"'},
    )
